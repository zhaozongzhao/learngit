from openpyxl import Workbook

#新建了一个新的工作表(没有保存)
wb = Workbook()
# 提供一个默认名叫Sheet的表
print(wb.sheetnames)
# 直接赋值就可以改工作表的名称
sheet =  wb['Sheet']
sheet.title = 'Sheet1'
print(wb.sheetnames)
# 新建一个工作表，可以指定索引，适当安排其在工作簿中的位置
#  被安排到第二个工作表，index=0就是第一个位置
wb.create_sheet('data',index=1)
print(wb.sheetnames)
#删除某个工作表
# wb.remove(wb['Sheet1'])
# print(wb.sheetnames)

print('*******************写入excel单元格*********************')
#写入单元格换还可以用公式
sheet['A1'] =  'good'
## B9处写入平均值
sheet['B9'] = '=AVERAGE(B2:B8)'
print(sheet['A1'].value)
print(sheet['B9'].value)

#append()函数,可以一次添加多行数据,从第一行空白行开始

#添加一行
row = [1,2,3,4,5]
sheet.append(row)
#添加多行
rows = [
    ['Number', 'data1', 'data2'],
    [2, 40, 30],
    [3, 40, 25],
    [4, 50, 30],
    [5, 30, 10],
    [6, 25, 5],
    [7, 50, 10],
]
sheet.append(row)

#由于append函数只能按行输入,如果我们想按列输入如果把上面的列表嵌套看作矩阵。
#  只要将矩阵转置就可以了。使用zip()函数可以实现，不过内部的列表变成了元组就是了。都是可迭代对象，不影
list(zip(*rows))

print(list(zip(*rows)))
sheet.append(row)

wb.save('工作簿2.xlsx')