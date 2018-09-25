from openpyxl import load_workbook

wd = load_workbook('/Users/hnbl009/Desktop/工作簿1.xlsx')

#获取所有的sheet名称
print(wd.get_sheet_names())#弃用
print(wd.sheetnames)

#根据sheet名称获取sheet
a_sheet = wd.get_sheet_by_name('工作表1')#弃用
a_sheet2 = wd['工作表1']

#获取sheet名称
print(a_sheet.title)
print(a_sheet2.title)

#获取当前正在显示的sheet,也可以用
print(wd.get_active_sheet())#弃用
print(wd.active)

#获取单元格
b4 = a_sheet2['B4']
#分别返回
print(b4.column,b4.row,b4.value)

#还可以用下标的方式获得,还可以用cell函数,换成数字
b4_too = a_sheet2.cell(row =4,column= 2)
print(b4_too)

#获取最大行和最大列
print(a_sheet2.max_row)
print(a_sheet2.max_column)


#获取行和列
for row  in a_sheet2.rows:
    for cell in row:
        print(cell.value)


for column in a_sheet.columns:
    for cell in column:
        print(cell.value)


#获取指定列的数据
for row  in a_sheet2.rows:
    for cell in list(row)[2]:
        print(cell.value)

print('*************************************')
#获取指定行的数据
for cell in list(a_sheet2.rows)[2]:
 print(cell.value)

#获取任意区间的单元格
for i in range(1, 4):
    for j in range(1, 3):
        print(a_sheet2.cell(row=i, column=j).value)

